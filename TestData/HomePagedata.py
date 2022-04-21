import openpyxl


class HomePagedata:

    test_homepage_data =  [{"Name":"Anton","Email":"ATN505","Gender":"Male"},{"Name":"Yana","Email":"ftw","Gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\DriversForSelenium\ExcelDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):  # 2 loops to print all columns and rows
            if sheet.cell(row=i, column=1).value == test_case_name:  # chekin if cell name is t3 on column 1
                for x in range(2, sheet.max_column + 1):  # put all the t3 row in dictionary
                    Dict[sheet.cell(row=1, column=x).value] = sheet.cell(row=i,column=x).value  # put row 1 column 1 in dictionary on first row column 1 and then do +1 on all
        print(Dict)
        return [Dict]