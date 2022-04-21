import openpyxl
book = openpyxl.load_workbook("C:\DriversForSelenium\ExcelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=1) # Read
print(cell.value)
#sheet.cell(row=2, column=2).value = "First one" # write into excel
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row) # Print max row
print(sheet.max_column) # print max columns
print(sheet['A6'].value) # another way to exract value
print("*************************************")

Dict= {}
for i in range(1,sheet.max_row+1): # 2 loops to print all columns and rows
    if sheet.cell(row=i,column=1).value == "TestCase2": # chekin if cell name is t3 on column 1
        for x in range(2,sheet.max_column+1): # put all the t3 row in dictionary
            Dict[sheet.cell(row=1,column=x).value]=sheet.cell(row=i,column=x).value # put row 1 column 1 in dictionary on first row column 1 and then do +1 on all
print(Dict)
